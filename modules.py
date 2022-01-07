import openpyxl as xl
import os
from FlightMember import FlightMember
from openpyxl.chart import LineChart, Reference


def processWorkbook(directory, wb, Filename):
    # Setup Output File
    directoryOut = directory + '\ProcessedData'
    OUTPUT_FILE = os.path.splitext(Filename)[0] + '_Processed' + os.path.splitext(Filename)[1]
    wbOut = xl.Workbook()  # Create output workbook to store info and rankings on
    fltSheet = wbOut.active
    fltSheet.title = "Flight Info"  # First sheet for flight info
    # Loop through all flight members and generate their analysis
    rawNames = wb['Sheet1']
    roster = []
    avgRanks = []
    selfRanks = []
    for col in range(2, rawNames.max_column + 1):  # Loop through columns of raw rankings
        cell = rawNames.cell(1, col)
        roster.append(cell.value)  # Adds name to roster
    for col in range(2, rawNames.max_column + 1):  # Loop through columns of raw rankings
        cell = rawNames.cell(1, col)
        member = FlightMember(rawNames, wbOut, cell.value, Filename)  # Establish flt member and preform rank processing
        avgRanks.append(member.avgRank)  # Adds corresponding avg rank to list
        selfRanks.append(member.selfRanking)
        # Check for missing names in ranking
        s = set(member.ranking)
        missing = [x for x in roster if x not in s]
        if missing:  # List is not empty
            print("[ERROR] ({}) Missing name from {}'s ranking: {}".format(member.filename, member.name, missing))
    # Order lists according to avg rank
    avgRanks_ordered, roster_ordered = orderSimultaneously(avgRanks, roster)
    # Add Flight Info
    flightInfo(fltSheet, avgRanks_ordered, roster_ordered)  # Pass ordered roster & ranks
    # Save workbook
    try:
        wbOut.save(os.path.join(directoryOut, OUTPUT_FILE))
    except FileNotFoundError:
        os.mkdir(os.path.join(directoryOut))
        wbOut.save(os.path.join(directoryOut, OUTPUT_FILE))  # Save workbook to ProcessedData

    return roster, avgRanks, selfRanks


def orderSimultaneously(list1, list2):  # Ordered according to list 1
    zipped_lists = zip(list1, list2)
    sorted_pairs = sorted(zipped_lists)

    tuples = zip(*sorted_pairs)
    list1, list2 = [list(tuple) for tuple in tuples]
    return list1, list2


def flightInfo(ws, ranks, roster):
    header = {
        'A1': "Order:",
        'B1': "Name:",
        'C1': "Avg Rank:"
    }
    # Number each member in roster
    for row in range(2, len(roster) + 2):
        ws.cell(column=1, row=row, value=(row - 1))
    # Add items in template to xlsx
    for item in header.items():
        ws[item[0]] = item[1]
    # Add ranking list to xlsx
    for row in range(2, len(roster) + 2):
        ws.cell(column=2, row=row, value=roster[row - 2])
        ws.cell(column=3, row=row, value=ranks[row - 2])


def createSummary(directory, fltRosters, fltAvgRanks, fltSelfRanks):
    directoryOut = directory + '\ProcessedData'
    OUTPUT_FILE = 'Summary.xlsx'
    wbSummary = xl.Workbook()  # Create output workbook to store info and rankings on
    fltSheet = wbSummary.active
    fltSheet.title = "Flight Info"  # First sheet for flight info
    count = 0
    for members, ranks, selfRanks in zip(fltRosters, fltAvgRanks, fltSelfRanks):
        count += 1
        for member, rank, selfRank in zip(members, ranks, selfRanks):
            fillSheet(wbSummary, member, rank, selfRank, count)

    # Save Summary workbook
    try:
        wbSummary.save(os.path.join(directoryOut, OUTPUT_FILE))
    except FileNotFoundError:
        os.mkdir(os.path.join(directoryOut))
        wbSummary.save(os.path.join(directoryOut, OUTPUT_FILE))  # Save workbook to ProcessedData


def fillSheet(wb, member, rank, selfRank, count):
    template = {
        'A1': "Name:",
        'A2': "Week #:",
        'B2': "Avg Rank:",
        'C2': "Self Rank:",
    }
    template.update({'B1': member})
    # Graph stuff
    chart = LineChart()
    chart.style = 10
    chart.title = "Ranking Tracker: {}".format(member)
    chart.y_axis.title = 'Rank'
    chart.x_axis.title = 'Week #'

    if member not in wb.sheetnames:  # Member does not have a sheet yet
        ws = wb.create_sheet(member)  # Create a worksheet for flight member
        # Add items in template to xlsx
        for item in template.items():
            ws[item[0]] = item[1]

        # Add data to table and graph
        addData(ws, count, rank, selfRank, chart)

    else:  # Member already has a sheet that should be added to
        ws = wb[member]  # Access their ws
        # Add data to table and graph
        addData(ws, count, rank, selfRank, chart)


def addData(ws, count, rank, selfRank, chart):
    # Add to table
    ws.cell(column=1, row=count + 2, value=count)  # Add week number
    ws.cell(column=2, row=count + 2, value=rank)  # Add avg rank
    ws.cell(column=3, row=count + 2, value=selfRank)  # Add self rank
    # Add to graph
    data = Reference(ws, min_col=2, min_row=2, max_col=3, max_row=count + 2)
    chart.add_data(data, titles_from_data=True)
    ws.add_chart(chart, 'E2')
