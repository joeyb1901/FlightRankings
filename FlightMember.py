from openpyxl.chart import BarChart, Reference

class FlightMember:
    def __init__(self, rawNames, outputWB, name, filename):
        self.filename = filename
        self.name = name
        self.sheet = outputWB.create_sheet(self.name)  # Create a worksheet for flight member
        self.rawNames = rawNames  # xlsx of names (row 1 is ranker, col 1 is numbers)
        self.ranking = []  # Member's ranking of others
        self.selfRanking = None  # How they rank themself
        self.getRanking()  # How they rank their flight
        self.findDuplicates()  # Check the ranking for duplicate names
        self.ROSTER_LENGTH = len(self.ranking)
        self.rank = []  # Others' ranking of the member
        self.getRank()
        self.avgRank = round(sum(self.rank) / len(self.rank), 2)
        self.fillSheet()  # Create sheet for member on output xlsx

    def getRanking(self):
        for col in range(2, self.rawNames.max_column + 1):  # Loop through columns
            cell = self.rawNames.cell(1, col)
            if self.name == cell.value:  # Find name of interest in top row
                index = col  # Column of interest
                for row in range(2, self.rawNames.max_row + 1):
                    cell = self.rawNames.cell(row, index)
                    self.ranking.append(cell.value)
                    if cell.value == self.name:  # Find self ranking
                        self.selfRanking = row - 1
                break  # Skips unnecessary columns

    def findDuplicates(self):
        seen = set()  # list of seen names in ranking
        dupes = [x for x in self.ranking if x in seen or seen.add(x)]
        if dupes:  # List is not empty
            print("[ERROR] ({}) Duplicates found in {}'s ranking: {}".format(self.filename, self.name, dupes))

    def getRank(self):
        # Loop through all ranking lists in rawNames to generate list of ranks for this member
        for col in range(2, self.rawNames.max_column + 1):
            for row in range(2, self.rawNames.max_row + 1):
                cell = self.rawNames.cell(row, col)
                if cell.value == self.name:
                    self.rank.append(row - 1)

    def fillSheet(self):
        template = {
            'A1': "Name:",
            'A2': "Avg Rank:",
            'A3': "Self Rank:",
            'A4': "Flight Ranking:",
            'C4': "Histogram:"
        }
        # Fill B column with corresponding information
        template.update({'B1': self.name})
        template.update({'B2': "{} of {}".format(self.avgRank, self.ROSTER_LENGTH)})
        template.update({'B3': "{} of {}".format(self.selfRanking, self.ROSTER_LENGTH)})
        # Add items in template to xlsx
        for item in template.items():
            self.sheet[item[0]] = item[1]
        # Add ranking list to xlsx
        for row in range(5, self.ROSTER_LENGTH + 5):
            self.sheet.cell(column=1, row=row, value=(row - 4))
            self.sheet.cell(column=2, row=row, value=self.ranking[row - 5])
        # Create and add histogram
        self.createHistogram()  # Inserted to 'F2'

    def createHistogram(self):
        # Create and label chart
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Peer Ranking Histogram: {}".format(self.name)
        chart.y_axis.title = 'Frequency'
        chart.x_axis.title = 'Rank among flight'
        # Generate Histogram List
        hist = [0] * self.ROSTER_LENGTH
        for i in range(0, len(self.rank)):
            hist[self.rank[i]-1] += 1
        # Add histogram list to the spreadsheet
        for row in range(5, self.ROSTER_LENGTH + 5):
            self.sheet.cell(column=3, row=row, value=(row - 4))
            self.sheet.cell(column=4, row=row, value=hist[row - 5])
        # Add data and labels to chart
        data = Reference(self.sheet, min_col=4, min_row=5, max_row=4+self.ROSTER_LENGTH)  # Location of hist on sheet
        chart.add_data(data, titles_from_data=False)
        labels = Reference(self.sheet, min_col=1, min_row=5, max_row=(4+len(self.ranking)))
        chart.set_categories(labels)
        # Add chart to xlsx
        self.sheet.add_chart(chart, 'F2')
